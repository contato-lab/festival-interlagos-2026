#!/usr/bin/env python3
"""
update-vendas-unificado.py
Festival Interlagos 2026 — Agência Lime

Script ÚNICO que substitui os três anteriores:
  - update-vendas-proprias.py  (vendas-data.json)
  - update-ticketmaster.py     (ticketmaster-data.json)
  - update-vendas-tipos.py     (vendas-tipos-data.json)

Ele faz uma única leitura de cada API (Sistema Próprio + Ticketmaster) e
gera os três arquivos JSON a partir do MESMO snapshot. Isso elimina a
defasagem entre os pipelines, que antes geravam números divergentes na
seção "Tipos de Ingresso" do dashboard.

Roda via GitHub Actions a cada 2 minutos.
"""

import io
import json
import os
import re
import sys
import requests
from collections import defaultdict
from datetime import date, datetime, timedelta, timezone

# ─── CONFIGURAÇÃO ──────────────────────────────────────────
API_MOTO_BASE   = 'https://ingressosmoto.festivalinterlagos.com.br'
API_AUTO_BASE   = 'https://ingressosauto.festivalinterlagos.com.br'
DATA_INICIO     = date(2026, 3, 31)
DATA_INICIO_STR = '2026-01-01'
PAGE_SIZE       = 100
# Dias rebuscados numa consulta curta e propria, pra o dia corrente nunca
# depender de a varredura longa (50+ paginas) chegar ate o fim.
# ─── O QUE CONTA COMO VENDA ────────────────────────────────
# O log de 16/08/2026 mostrou tres status vindos da API do proprio:
#   3                 -> venda paga
#   finalizado-manual -> venda fechada na mao (bilheteria, ajuste). E VENDA.
#   cupom             -> cortesia de valor zero. NAO e venda.
# O robo so contava '3', entao jogava fora as finalizadas manualmente: 15 hoje,
# 15 ontem, 8 no dia 14. Cupom fica de fora de proposito, senao o faturamento
# do painel passa a contar 1.186 entradas gratuitas de hoje como se fossem
# venda.
STATUS_VENDA = ('3', 'finalizado-manual')
RECENTE_DIAS    = 3
# Pagina grande so nas consultas curtas (dia de hoje e vizinhanca). A varredura
# longa continua em 100: pedir 500 nela aumenta a chance de a API engasgar.
PAGE_SIZE_DIA   = 500

# API v2.0 (a partir de 29/07/2026): o token deixou de ser publico.
# Agora e POST /apis/token com a chave no header X-Api-Key, e o token vale 1 HORA.
# A chave NUNCA entra no codigo: este repositorio e publico.
FI_API_KEY = os.environ.get('FI_API_KEY', '')

TM_API_BASE       = 'https://data.getcrowder.com'
TM_API_ENDPOINT   = '/activity/organizer'
TM_API_KEY        = os.environ.get('TM_API_KEY', '52087b883f65e8d2a684ee680c5beca66e425fdcc46c2b653da0fffd50088734')
CAMPAIGN_START_MS = 1774396800000  # 2026-03-25 00:00:00 UTC
MOTO_SHOW_IDS = {195330, 195736, 195737, 195738}
AUTO_SHOW_IDS = {195331, 195739, 195740, 195741}
SHOW_TO_DATE = {
    195330: '13/08/2026', 195736: '14/08/2026', 195737: '15/08/2026', 195738: '16/08/2026',
    195331: '27/08/2026', 195739: '28/08/2026', 195740: '29/08/2026', 195741: '30/08/2026',
}

OUT_VENDAS = 'vendas-data.json'
OUT_TM     = 'ticketmaster-data.json'
OUT_TIPOS  = 'vendas-tipos-data.json'
OUT_HORA   = 'vendas-por-hora.json'

# ── PLANILHA (fallback quando a API TM cai) ────────────────
PLANILHA_ID  = os.environ.get('PLANILHA_LINHA_TEMPO_ID', '1osgyQdQlf21kbsZCt65CnrbsONTdiLyXAd32Hngvbfk')
PLANILHA_URL = f'https://docs.google.com/spreadsheets/d/{PLANILHA_ID}/export?format=xlsx'
PLANILHA_YEAR = 2026
# SEMANA 0 começa nessa segunda-feira (data fixa da campanha).
# SEMANA N = SEMANA 0 + N*7 dias.
PLANILHA_SEMANA_0_START = date(2026, 3, 30)


# ╔══════════════════════════════════════════════════════════╗
# ║  SISTEMA PRÓPRIO                                        ║
# ╚══════════════════════════════════════════════════════════╝

def _proprio_headers(base, token=None):
    h = {
        'Accept':           'application/json',
        'Content-Type':     'application/json',
        'Origin':           base,
        'Referer':          base + '/',
        'User-Agent':       'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 '
                            '(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36',
        'X-Requested-With': 'XMLHttpRequest',
    }
    if token:
        h['Authorization'] = f'Bearer {token}'
    return h


def get_proprio_token(base):
    """API v2.0: troca a chave por um token de 1 hora. POST + header X-Api-Key."""
    if not FI_API_KEY:
        raise RuntimeError(
            'FI_API_KEY nao definida. A API do sistema proprio virou v2.0 em 29/07/2026 '
            'e exige a chave do Festival Interlagos. Cadastre o secret FI_API_KEY no '
            'repositorio (Settings > Secrets and variables > Actions).')
    h = _proprio_headers(base)
    h['X-Api-Key'] = FI_API_KEY
    r = requests.post(base + '/apis/token', headers=h, timeout=30)
    r.raise_for_status()
    d = r.json()
    if d.get('status') != 'success' or not d.get('token'):
        raise RuntimeError(f'/apis/token nao devolveu token: {str(d)[:200]}')
    return d['token']


def fetch_proprio_sales(base):
    """Busca todas as vendas do Sistema Próprio, renovando o token se expirar."""
    token = get_proprio_token(base)

    fim = datetime.now().strftime('%Y-%m-%d')
    sales = _paginar_vendas(base, token, DATA_INICIO_STR, fim)

    # ── REBUSCA DOS ULTIMOS DIAS, SEPARADO ──────────────────────────
    # A consulta longa varre desde 01/01 em paginas de 100, o que hoje passa de
    # 50 paginas. Se UMA pagina falhar no meio, a paginacao para e a lista
    # parcial passa por completa. Como a API devolve da mais antiga pra mais
    # nova, o que se perde e sempre o FINAL: o dia corrente.
    #
    # Foi exatamente o que aconteceu em 16/08/2026: o admin mostrava 6 vendas
    # aprovadas no dia e o painel mostrava zero, porque o dia nunca chegava.
    #
    # Esta segunda consulta cobre so os ultimos dias. Cabe numa pagina ou duas,
    # entao nao depende de paginacao longa dar certo. O merge por chave da
    # venda evita contar duas vezes quando as duas consultas trazem o mesmo
    # registro.
    try:
        desde = (date.today() - timedelta(days=RECENTE_DIAS)).strftime('%Y-%m-%d')
        recentes = _paginar_vendas(base, token, desde, fim)
        if recentes:
            vistos = {_chave_venda(v) for v in sales}
            novas = [v for v in recentes if _chave_venda(v) not in vistos]
            if novas:
                print(f'  + {len(novas)} vendas recentes que a varredura longa nao trouxe')
            sales.extend(novas)
    except Exception as e:
        # a rebusca e reforco: se falhar, segue com o que a varredura longa deu
        print(f'  aviso: rebusca dos ultimos dias falhou ({e})', file=sys.stderr)

    return sales


def fetch_proprio_sales_resiliente(base, rotulo):
    """Busca as vendas SEM depender da varredura longa dar certo.

    A varredura desde 01/01 tem mais de 50 paginas e a API interrompe no meio
    (comprovado em 16/08/2026: a protecao nova derrubou a rodada #1477). Como a
    resposta vem da venda mais antiga pra mais nova, quando ela quebra o que se
    perde e sempre o FINAL, ou seja, o dia corrente. Era por isso que o painel
    mostrava zero num dia com venda no admin.

    Agora a janela recente e buscada SOZINHA, numa consulta curta de uma ou
    duas paginas que nao depende da longa. Se a longa falhar, o historico e
    reaproveitado do arquivo anterior, que foi montado quando ela deu certo.

    Devolve (vendas, historico_confiavel).
    """
    token = get_proprio_token(base)
    fim = datetime.now().strftime('%Y-%m-%d')

    # 1) O DIA DE HOJE, SOZINHO E EM PAGINA GRANDE.
    #    Minha primeira tentativa usou janela de 3 dias, que durante o evento
    #    cobre 13 a 16/08 e tem dezenas de paginas: quebrava igual a varredura
    #    longa e trouxe 1 venda quando o admin mostrava 6. Um dia so, com
    #    pagina de 500, cabe numa requisicao e nao depende de paginacao.
    hoje = date.today().strftime('%Y-%m-%d')
    recentes = _paginar_vendas(base, token, hoje, fim, page_size=PAGE_SIZE_DIA)
    print(f'  [{rotulo}] hoje ({hoje}): {len(recentes)} vendas')

    # 2) Os dias em volta, pra pegar venda que mudou de status depois.
    #    Se falhar, o dia de hoje ja esta garantido acima.
    try:
        desde = (date.today() - timedelta(days=RECENTE_DIAS)).strftime('%Y-%m-%d')
        vizinhos = _paginar_vendas(base, token, desde, fim, page_size=PAGE_SIZE_DIA)
        vistos_h = {_chave_venda(v) for v in recentes}
        recentes += [v for v in vizinhos if _chave_venda(v) not in vistos_h]
        print(f'  [{rotulo}] janela {desde}..{fim}: {len(recentes)} vendas no total')
    except Exception as e:
        print(f'  [{rotulo}] janela de {RECENTE_DIAS} dias falhou ({e}). '
              f'Hoje segue garantido.', file=sys.stderr)

    # 2) HISTORICO. Se falhar, nao derruba a rodada: o dia de hoje ja esta salvo.
    historico_ok = True
    try:
        historico = _paginar_vendas(base, token, DATA_INICIO_STR, fim)
        print(f'  [{rotulo}] varredura longa: {len(historico)} vendas')
    except Exception as e:
        historico, historico_ok = [], False
        print(f'  [{rotulo}] varredura longa FALHOU ({e}). '
              f'Historico vem do arquivo anterior.', file=sys.stderr)

    vistos = {_chave_venda(v) for v in recentes}
    todas = recentes + [v for v in historico if _chave_venda(v) not in vistos]
    return todas, historico_ok


def _dia_de_corte():
    """Primeiro dia (numero D+) que a janela recente cobre."""
    return (date.today() - timedelta(days=RECENTE_DIAS) - DATA_INICIO).days + 1


def _historico_do_arquivo(campo_receita, campo_qtd):
    """Dias antigos vindos do vendas-data.json anterior, pra quando a varredura
    longa falha. Melhor um historico de minutos atras que um zero inventado."""
    prev = _load_existing(OUT_VENDAS) or {}
    out = {}
    for row in (prev.get('daily') or []):
        d = row.get('d')
        if isinstance(d, int):
            out[d] = {'receita': float(row.get(campo_receita) or 0),
                      'qtd': int(row.get(campo_qtd) or 0)}
    return out


def mesclar_historico(by_day, historico_ok, campo_receita, campo_qtd, rotulo):
    """Quando a varredura longa falha, os dias ANTIGOS vem do arquivo anterior e
    os dias da janela recente vem da API. Assim uma falha no historico nunca
    apaga o dia de hoje, nem o dia de hoje zera o historico."""
    if historico_ok:
        return by_day
    corte = _dia_de_corte()
    base = _historico_do_arquivo(campo_receita, campo_qtd)
    juntos = {d: v for d, v in base.items() if d < corte}
    for d, v in by_day.items():
        juntos[d] = v
    print(f'  [{rotulo}] historico reaproveitado do arquivo ate D+{corte - 1}, '
          f'janela recente da API a partir de D+{corte}')
    return juntos


def _chave_venda(v):
    """Identidade da venda pro merge. Nao existe campo de id exposto, entao uso
    a combinacao que nao se repete na pratica: momento da compra, valor e os
    proprios qrcodes."""
    qr = v.get('qrcodes') or []
    marca = ','.join(sorted(str(q.get('codigo') or q.get('id') or q) for q in qr)) if qr else ''
    return (str(v.get('created_at') or ''), str(v.get('venda_valor') or ''), marca)


def _paginar_vendas(base, token, dt_inicio, dt_fim, page_size=None):
    """Percorre as paginas de /apis/vendas. LANCA se a paginacao morrer no meio:
    devolver lista parcial como se fosse completa e o que fez o dashboard
    mostrar zero num dia com venda."""
    out, page, renovou = [], 1, False
    ps = page_size or PAGE_SIZE
    while True:
        url = (f'{base}/apis/vendas?data_inicio={dt_inicio}&data_fim={dt_fim}'
               f'&page_size={ps}&page={page}')
        r = requests.get(url, headers=_proprio_headers(base, token), timeout=30)
        # token de 1h pode vencer no meio da paginacao: renova uma vez e repete a pagina
        if r.status_code == 401 and not renovou:
            renovou = True
            token = get_proprio_token(base)
            continue
        r.raise_for_status()
        renovou = False
        d = r.json()
        if d.get('status') != 'success':
            # pagina 1 sem sucesso = janela vazia de verdade. Da pagina 2 em
            # diante e falha no meio, e ai nao da pra fingir que acabou.
            if page == 1:
                return out
            raise RuntimeError(f'paginacao interrompida na pagina {page} '
                               f'({dt_inicio}..{dt_fim}): status={d.get("status")}')
        if not d.get('data'):
            return out
        out.extend(d['data'])
        if d.get('pagination', {}).get('isNextPage') != 'Y':
            return out
        page += 1


def diagnostico_recentes(sales, rotulo):
    """Conta o que a API devolveu nos ultimos dias, POR STATUS.

    Existe porque em 16/08/2026 o admin mostrava 6 vendas aprovadas no dia e o
    dia nao aparecia no painel. Sem ver status e data crus nao da pra saber se
    a venda nao vem, se vem com outro status, ou se vem com outra data. Deduzir
    isso de fora ja custou uma manha.
    """
    corte = (date.today() - timedelta(days=3)).isoformat()
    por_dia = {}
    maior = ''
    for v in sales:
        ds = (v.get('created_at') or '')[:10]
        if ds > maior:
            maior = ds
        if ds >= corte:
            st = str(v.get('venda_status'))
            por_dia.setdefault(ds, {}).setdefault(st, 0)
            por_dia[ds][st] += 1
    print(f'  [diag {rotulo}] created_at mais recente que a API devolveu: {maior or "nenhum"}')
    for ds in sorted(por_dia):
        detalhe = ', '.join(f'status {k}: {n}' for k, n in sorted(por_dia[ds].items()))
        print(f'  [diag {rotulo}] {ds} -> {detalhe}')
    if not por_dia:
        print(f'  [diag {rotulo}] NENHUMA venda nos ultimos 3 dias veio da API')


def aggregate_proprio_totais(sales):
    """Agrega para vendas-data.json (totais diários: receita + qtd ingressos)."""
    by_day = {}
    for v in sales:
        if str(v.get('venda_status')) not in STATUS_VENDA:
            continue
        ds = (v.get('created_at') or '').split(' ')[0]
        if not ds or len(ds) < 10:
            continue
        try:
            dt = date.fromisoformat(ds)
        except ValueError:
            continue
        d = (dt - DATA_INICIO).days + 1
        if d < 1:
            continue
        if d not in by_day:
            by_day[d] = {'receita': 0.0, 'qtd': 0, 'date': ds}
        by_day[d]['receita'] += float(v.get('venda_valor') or 0)
        by_day[d]['qtd']     += len(v.get('qrcodes') or [])
    return by_day


def aggregate_proprio_tipos(sales):
    """Agrega para vendas-tipos-data.json (breakdown por ingresso_nome)."""
    def make():
        return {
            'qtd': 0, 'rec': 0.0,
            'daily': defaultdict(lambda: {'qtd': 0, 'rec': 0.0}),
            'por_dia_evento': defaultdict(lambda: {
                'qtd': 0, 'rec': 0.0,
                'daily': defaultdict(lambda: {'qtd': 0, 'rec': 0.0})
            }),
        }
    agg = defaultdict(make)
    for v in sales:
        if str(v.get('venda_status')) not in STATUS_VENDA:
            continue
        ds = (v.get('created_at') or '').split(' ')[0]
        if not ds or len(ds) < 10:
            continue
        for q in (v.get('qrcodes') or []):
            nome    = q.get('ingresso_nome', 'Desconhecido')
            valor   = float(q.get('ingresso_valor') or 0)
            data_ev = q.get('data', '') or 'Sem data'
            agg[nome]['qtd'] += 1
            agg[nome]['rec'] += valor
            agg[nome]['daily'][ds]['qtd'] += 1
            agg[nome]['daily'][ds]['rec'] += valor
            de = agg[nome]['por_dia_evento'][data_ev]
            de['qtd'] += 1
            de['rec'] += valor
            de['daily'][ds]['qtd'] += 1
            de['daily'][ds]['rec'] += valor
    out = {}
    for nome, d in agg.items():
        out[nome] = {
            'qtd': d['qtd'],
            'rec': round(d['rec'], 2),
            'daily': {ds: {'qtd': dd['qtd'], 'rec': round(dd['rec'], 2)}
                      for ds, dd in d['daily'].items()},
            'por_dia_evento': {de: {
                'qtd': dde['qtd'],
                'rec': round(dde['rec'], 2),
                'daily': {ds: {'qtd': sd['qtd'], 'rec': round(sd['rec'], 2)}
                          for ds, sd in dde['daily'].items()}
            } for de, dde in d['por_dia_evento'].items()},
        }
    return out


# ╔══════════════════════════════════════════════════════════╗
# ║  TICKETMASTER                                            ║
# ╚══════════════════════════════════════════════════════════╝

def fetch_tm_movements():
    """Busca todos os movimentos do Ticketmaster. Uma única vez.
    Retorna tambem o cursor final (lastUpdate/lastMovementId) pra o Worker
    de tempo real conseguir retomar dai em diante (busca incremental)."""
    all_movs = []
    last_update, last_mov_id = CAMPAIGN_START_MS, 1
    while True:
        url = (f'{TM_API_BASE}{TM_API_ENDPOINT}'
               f'?lastUpdate={last_update}&lastMovementId={last_mov_id}')
        r = requests.get(url, headers={'apiKey': TM_API_KEY,
                                       'Content-Type': 'application/json'},
                         timeout=30)
        r.raise_for_status()
        d = r.json()
        movs = d.get('movements', [])
        all_movs.extend(movs)
        # Avanca o cursor SEMPRE, inclusive na ultima pagina. Antes o break vinha
        # antes desta atualizacao, entao o cursor ficava posicionado ANTES da ultima
        # pagina de movimentos: o Worker incremental re-buscava essa pagina e contava
        # tudo em dobro (inflou o dia atual em +32 em 03/07/2026).
        last_update = d.get('lastUpdate', last_update)
        last_mov_id = d.get('lastMovementId', last_mov_id)
        if not d.get('hasMore') or not movs:
            break
    return all_movs, last_update, last_mov_id


def _classify_show(tickets):
    for t in tickets:
        sid = t.get('show', {}).get('id')
        if sid in MOTO_SHOW_IDS:
            return 'moto'
        if sid in AUTO_SHOW_IDS:
            return 'auto'
    return None


def _classify_meia_inteira(rate_name, rate_cat):
    cat  = (rate_cat or '').lower()
    nome = (rate_name or '').lower()
    if 'meia' in cat or 'meia' in nome or 'estatuto idoso' in nome:
        return 'meia'
    return 'inteira'


def aggregate_tm_totais(movements):
    """Agrega para ticketmaster-data.json (totais diários por edição)."""
    issuance_date = {}
    cortesia_purchase = set()  # ids de compra cuja emissao original foi cortesia (amt=0)
    for m in movements:
        if m.get('operation') == 'ISSUANCE':
            pid = (m.get('purchase') or {}).get('id')
            if pid is not None and pid not in issuance_date:
                issuance_date[pid] = (m.get('date') or '')[:10]
            if pid is not None and float(m.get('amount', 0)) == 0 and int(m.get('ticketCount', 0)) > 0:
                cortesia_purchase.add(pid)

    daily = defaultdict(lambda: {
        'moto_receita': 0.0, 'auto_receita': 0.0,
        'moto_ingressos': 0,  'auto_ingressos': 0,
        'moto_cortesias': 0,  'auto_cortesias': 0,
    })
    totals = {
        'moto_receita': 0.0, 'auto_receita': 0.0,
        'moto_ingressos': 0,  'auto_ingressos': 0,
        'moto_cortesias': 0,  'auto_cortesias': 0,
    }
    last_sale_moto_at = None
    last_sale_auto_at = None

    for m in movements:
        op   = m.get('operation')
        amt  = float(m.get('amount', 0))
        tc   = int(m.get('ticketCount', 0))
        edi  = _classify_show(m.get('tickets', []))
        if not edi:
            continue
        pid = (m.get('purchase') or {}).get('id')
        if op in ('CANCELLATION', 'REFUND'):
            ds  = issuance_date.get(pid) or (m.get('date') or '')[:10]
        else:
            ds = (m.get('date') or '')[:10]
        if not ds or len(ds) < 10:
            continue

        # Cortesias / vouchers: ISSUANCE com R$ 0 não são vendas comerciais.
        # Em 21/05/2026 entraram 151 cortesias de moto inflando o número
        # de "ingressos vendidos". Separamos em moto_cortesias/auto_cortesias.
        # Cancelamento/reembolso de uma compra que NASCEU cortesia tem que abater
        # da MESMA conta (senao ele cai nas vendas reais e derruba o numero de
        # ingressos vendidos por engano — bug real observado em 01/07/2026,
        # moto_ingressos foi parar negativo).
        if op == 'ISSUANCE':
            is_cortesia = (amt == 0 and tc > 0)
        elif op in ('CANCELLATION', 'REFUND'):
            is_cortesia = pid in cortesia_purchase
        else:
            is_cortesia = False

        if edi == 'moto':
            if is_cortesia:
                daily[ds]['moto_cortesias'] += tc
                totals['moto_cortesias']    += tc
            else:
                daily[ds]['moto_receita']   += amt
                daily[ds]['moto_ingressos'] += tc
                totals['moto_receita']      += amt
                totals['moto_ingressos']    += tc
                if op == 'ISSUANCE' and tc > 0:
                    mv_date = m.get('date') or ''
                    if mv_date and (last_sale_moto_at is None or mv_date > last_sale_moto_at):
                        last_sale_moto_at = mv_date
        else:
            if is_cortesia:
                daily[ds]['auto_cortesias'] += tc
                totals['auto_cortesias']    += tc
            else:
                daily[ds]['auto_receita']   += amt
                daily[ds]['auto_ingressos'] += tc
                totals['auto_receita']      += amt
                totals['auto_ingressos']    += tc
                if op == 'ISSUANCE' and tc > 0:
                    mv_date = m.get('date') or ''
                    if mv_date and (last_sale_auto_at is None or mv_date > last_sale_auto_at):
                        last_sale_auto_at = mv_date

    totals['total_receita']   = round(totals['moto_receita'] + totals['auto_receita'], 2)
    totals['total_ingressos'] = totals['moto_ingressos'] + totals['auto_ingressos']
    totals['moto_receita']    = round(totals['moto_receita'], 2)
    totals['auto_receita']    = round(totals['auto_receita'], 2)

    daily_list = [{
        'date': ds,
        'moto_receita':   round(d['moto_receita'], 2),
        'auto_receita':   round(d['auto_receita'], 2),
        'moto_ingressos': d['moto_ingressos'],
        'auto_ingressos': d['auto_ingressos'],
        'moto_cortesias': d['moto_cortesias'],
        'auto_cortesias': d['auto_cortesias'],
    } for ds, d in sorted(daily.items())]

    return totals, daily_list, last_sale_moto_at, last_sale_auto_at


def aggregate_tm_tipos(movements):
    """Agrega para vendas-tipos-data.json (breakdown por product.name + meia/inteira).

    So VENDA entra aqui: cortesia (ISSUANCE R$ 0) e o cancelamento dela ficam de
    fora, senao o cancelamento cairia nas vendas e deixaria a qtd negativa (mesmo
    bug tratado em aggregate_tm_totais). A cortesia digital e mostrada, separada,
    no card de Cortesias."""
    issuance_date = {}
    cortesia_purchase = set()  # compras cuja emissao original foi cortesia (amt=0)
    for m in movements:
        if m.get('operation') == 'ISSUANCE':
            pid = (m.get('purchase') or {}).get('id')
            if pid is not None and pid not in issuance_date:
                issuance_date[pid] = (m.get('date') or '')[:10]
            if pid is not None and float(m.get('amount', 0)) == 0 and int(m.get('ticketCount', 0)) > 0:
                cortesia_purchase.add(pid)

    def make():
        return {
            'qtd': 0, 'rec': 0.0,
            'daily': defaultdict(lambda: {'qtd': 0, 'rec': 0.0}),
            'por_dia_evento': defaultdict(lambda: {
                'qtd': 0, 'rec': 0.0,
                'daily': defaultdict(lambda: {'qtd': 0, 'rec': 0.0})
            }),
            'breakdown': {
                'meia':    {'qtd': 0, 'rec': 0.0, 'daily': defaultdict(lambda: {'qtd': 0, 'rec': 0.0})},
                'inteira': {'qtd': 0, 'rec': 0.0, 'daily': defaultdict(lambda: {'qtd': 0, 'rec': 0.0})},
            },
        }
    moto = defaultdict(make)
    auto = defaultdict(make)

    for m in movements:
        edi = _classify_show(m.get('tickets', []))
        if not edi:
            continue
        prod = (m.get('product') or {}).get('name', '')
        if not prod:
            for t in m.get('tickets', []):
                prod = (t.get('sector') or {}).get('name', '')
                if prod:
                    break
        if not prod:
            prod = 'Desconhecido'

        op = m.get('operation')
        if op in ('CANCELLATION', 'REFUND'):
            pid = (m.get('purchase') or {}).get('id')
            ds  = issuance_date.get(pid) or (m.get('date') or '')[:10]
        else:
            ds = (m.get('date') or '')[:10]
        if not ds or len(ds) < 10:
            continue

        amt       = float(m.get('amount', 0))
        qtd       = int(m.get('ticketCount', 0))

        # Cortesia fora dos Tipos (so venda). Emissao cortesia = R$ 0; o
        # cancelamento dela pertence a uma compra que nasceu cortesia (pid no
        # set), entao tambem sai — assim nao vira -qtd nas vendas.
        pid = (m.get('purchase') or {}).get('id')
        if op == 'ISSUANCE':
            is_cortesia = (amt == 0 and qtd > 0)
        elif op in ('CANCELLATION', 'REFUND'):
            is_cortesia = pid in cortesia_purchase
        else:
            is_cortesia = False
        if is_cortesia:
            continue

        rate_name = (m.get('rate') or {}).get('name', '')
        rate_cat  = ((m.get('rate') or {}).get('category') or {}).get('name', '')
        tipo_mi   = _classify_meia_inteira(rate_name, rate_cat)

        data_ev = 'Sem data'
        for t in m.get('tickets', []):
            sid = t.get('show', {}).get('id')
            if sid in SHOW_TO_DATE:
                data_ev = SHOW_TO_DATE[sid]
                break

        bucket = moto if edi == 'moto' else auto
        b = bucket[prod]
        b['qtd'] += qtd
        b['rec'] += amt
        b['daily'][ds]['qtd'] += qtd
        b['daily'][ds]['rec'] += amt
        bd = b['breakdown'][tipo_mi]
        bd['qtd'] += qtd
        bd['rec'] += amt
        bd['daily'][ds]['qtd'] += qtd
        bd['daily'][ds]['rec'] += amt
        de = b['por_dia_evento'][data_ev]
        de['qtd'] += qtd
        de['rec'] += amt
        de['daily'][ds]['qtd'] += qtd
        de['daily'][ds]['rec'] += amt

    def to_dict(agg):
        out = {}
        for nome, d in agg.items():
            out[nome] = {
                'qtd': d['qtd'],
                'rec': round(d['rec'], 2),
                'daily': {ds: {'qtd': dd['qtd'], 'rec': round(dd['rec'], 2)}
                          for ds, dd in d['daily'].items()},
                'por_dia_evento': {de: {
                    'qtd': dde['qtd'],
                    'rec': round(dde['rec'], 2),
                    'daily': {ds: {'qtd': sd['qtd'], 'rec': round(sd['rec'], 2)}
                              for ds, sd in dde['daily'].items()}
                } for de, dde in d['por_dia_evento'].items()},
                'breakdown': {
                    mi: {
                        'qtd': bd['qtd'],
                        'rec': round(bd['rec'], 2),
                        'daily': {ds: {'qtd': dd['qtd'], 'rec': round(dd['rec'], 2)}
                                  for ds, dd in bd['daily'].items()},
                    } for mi, bd in d['breakdown'].items()
                },
            }
        return out

    return to_dict(moto), to_dict(auto)


# ╔══════════════════════════════════════════════════════════╗
# ║  ESCRITA DOS 3 JSONS                                     ║
# ╚══════════════════════════════════════════════════════════╝

def build_vendas_data(moto_by_day, auto_by_day):
    all_days = sorted(set(list(moto_by_day.keys()) + list(auto_by_day.keys())))
    daily, totals = [], {
        'moto_receita': 0.0, 'auto_receita': 0.0,
        'moto_ingressos': 0,  'auto_ingressos': 0,
    }
    for d in all_days:
        m = moto_by_day.get(d, {'receita': 0.0, 'qtd': 0})
        a = auto_by_day.get(d, {'receita': 0.0, 'qtd': 0})
        day_date = (DATA_INICIO + timedelta(days=d - 1)).isoformat()
        daily.append({
            'd': d, 'date': day_date,
            'moto_receita':   round(m['receita'], 2),
            'moto_ingressos': m['qtd'],
            'auto_receita':   round(a['receita'], 2),
            'auto_ingressos': a['qtd'],
        })
        totals['moto_receita']   += m['receita']
        totals['moto_ingressos'] += m['qtd']
        totals['auto_receita']   += a['receita']
        totals['auto_ingressos'] += a['qtd']
    totals['moto_receita']    = round(totals['moto_receita'], 2)
    totals['auto_receita']    = round(totals['auto_receita'], 2)
    totals['total_receita']   = round(totals['moto_receita'] + totals['auto_receita'], 2)
    totals['total_ingressos'] = totals['moto_ingressos'] + totals['auto_ingressos']

    return {
        'updated_at':     datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ'),
        'source':         'Plataforma própria (ingressosmoto/ingressosauto.festivalinterlagos.com.br)',
        'campaign_start': DATA_INICIO.isoformat(),
        'totals':         totals,
        'daily':          daily,
    }


def build_ticketmaster_data(totals, daily, source='api', last_sale_moto_at=None, last_sale_auto_at=None,
                            tm_cursor_last_update=None, tm_cursor_last_movement_id=None):
    source_label = {
        'api':      'Ticketmaster API via getcrowder.com',
        'planilha': 'Planilha Linha do Tempo (fallback enquanto API TM está fora)',
    }.get(source, 'Ticketmaster API via getcrowder.com')
    return {
        'updated_at':     datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ'),
        'source':         source_label,
        'source_type':    source,
        'campaign_start': '2026-03-25',
        'moto_show_ids':  sorted(MOTO_SHOW_IDS),
        'auto_show_ids':  sorted(AUTO_SHOW_IDS),
        'totals':         totals,
        'daily':          daily,
        'last_sale_moto_at': last_sale_moto_at,
        'last_sale_auto_at': last_sale_auto_at,
        # Cursor de paginacao onde essa rodada parou (usado pelo Worker de
        # tempo real pra buscar so o que e NOVO desde aqui, em vez de
        # reprocessar o historico inteiro a cada 5 min).
        'tm_cursor_last_update':     tm_cursor_last_update,
        'tm_cursor_last_movement_id': tm_cursor_last_movement_id,
    }


def build_tipos_data(proprio_moto, proprio_auto, tm_moto, tm_auto):
    return {
        'updated_at': datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ'),
        'source':     'Sistema Proprio + Ticketmaster API (snapshot unificado)',
        'proprio': {'moto': proprio_moto, 'auto': proprio_auto},
        'ticketmaster': {'moto': tm_moto, 'auto': tm_auto},
    }


def write_json(path, data):
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


# ╔══════════════════════════════════════════════════════════╗
# ║  PLANILHA (fallback quando API TM cai)                  ║
# ╚══════════════════════════════════════════════════════════╝

def _money_to_float(v):
    """Converte 'R$ 1.234,56' ou número direto para float. Vazio -> 0."""
    if v is None: return 0.0
    if isinstance(v, (int, float)): return float(v)
    s = str(v).strip()
    if not s or s in ('-', 'R$', 'R$ 0,00', 'R$ 0'): return 0.0
    s = s.replace('R$', '').replace(' ', '').strip()
    # formato BR: 1.234,56 -> 1234.56
    if ',' in s and s.rfind(',') > s.rfind('.'):
        s = s.replace('.', '').replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return 0.0


def _parse_qtd(v):
    if v is None: return 0
    if isinstance(v, (int, float)): return int(v)
    s = str(v).strip().replace('.', '').replace(',', '')
    if not s or s == '-': return 0
    try:
        return int(float(s))
    except ValueError:
        return 0


def _parse_week_header(header_text, sheet_name=None):
    """Determina as 7 datas ISO da semana. Estratégia:
    1) Se o NOME da aba é 'SEMANA N', usa SEMANA_0_START + N*7 dias (mais
       confiável — é como a planilha foi estruturada desde o início).
    2) Senão, tenta parsear 'DATA: DD A DD/MM' ou 'DATA: DD/MM A DD/MM'
       do cabeçalho (texto livre, sujeito a erro de digitação).
    """
    # 1) Sheet name
    if sheet_name:
        m = re.search(r'SEMANA\s+(\d+)', sheet_name.upper())
        if m:
            week_num = int(m.group(1))
            start = PLANILHA_SEMANA_0_START + timedelta(weeks=week_num)
            return [(start + timedelta(days=i)).isoformat() for i in range(7)]

    # 2) Header text fallback
    if not header_text:
        return None
    m = re.search(r'(\d{1,2})(?:/(\d{1,2}))?\s*[Aa]\s*(\d{1,2})/(\d{1,2})', header_text)
    if not m:
        return None
    try:
        end_d = int(m.group(3)); end_m = int(m.group(4))
        end_date = date(PLANILHA_YEAR, end_m, end_d)
    except ValueError:
        return None
    return [(end_date - timedelta(days=6 - i)).isoformat() for i in range(7)]


def fetch_planilha_data():
    """Baixa o XLSX e parseia as semanas. Devolve {tm: {moto, auto},
    proprio: {moto, auto}} no formato {date_iso: {qtd, rec}}. Lança
    exceção se download/parse falhar — caller decide o fallback."""
    from openpyxl import load_workbook
    r = requests.get(PLANILHA_URL, timeout=30)
    r.raise_for_status()
    wb = load_workbook(io.BytesIO(r.content), data_only=True, read_only=True)

    out = {
        'tm':      {'moto': {}, 'auto': {}},
        'proprio': {'moto': {}, 'auto': {}},
    }

    for sheet_name in wb.sheetnames:
        if not sheet_name.upper().startswith('SEMANA'):
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Cabeçalho com as datas (linha 0)
        header = rows[0]
        header_cells = [str(c) if c is not None else '' for c in header]
        header_text = ' '.join(header_cells)
        week_dates = _parse_week_header(header_text, sheet_name=sheet_name)
        if not week_dates:
            print(f'  [planilha] {sheet_name}: cabeçalho não reconhecido: "{header_text[:60]}"', file=sys.stderr)
            continue

        # Encontra cada bloco e suas linhas de VENDAS VALOR / QTD INGRESSOS
        # Estrutura: VENDAS MOTOS -> SISTEMA PRÓPRIO -> VENDAS VALOR -> QTD -> TICKET MASTER -> VENDAS VALOR -> QTD
        # Depois: VENDAS AUTOS -> mesmo pattern
        section = None       # 'moto' | 'auto' | None
        canal = None         # 'proprio' | 'tm' | None
        for row in rows[1:]:
            cells = [str(c).strip() if c is not None else '' for c in row]
            label = cells[0].upper() if cells else ''
            if not label:
                continue
            if 'VENDAS MOTOS' in label:
                section, canal = 'moto', None
                continue
            if 'VENDAS AUTOS' in label or 'VENDAS CARROS' in label:
                section, canal = 'auto', None
                continue
            if 'SISTEMA PR' in label:  # PRÓPRIO/PROPRIO
                canal = 'proprio'; continue
            if 'TICKET MASTER' in label or 'TICKETMASTER' in label:
                canal = 'tm'; continue
            # Para evitar capturar TOTAL VENDAS VALOR / TKT MÉDIO / BALANÇO / etc:
            if label.startswith('TOTAL') or 'TKT' in label or 'BALAN' in label or 'INVEST' in label:
                canal = None; continue

            if section and canal in ('proprio', 'tm'):
                if label.startswith('VENDAS VALOR'):
                    for i, day_iso in enumerate(week_dates):
                        v = _money_to_float(row[i + 1]) if i + 1 < len(row) else 0.0
                        if v == 0 and day_iso not in out[canal][section]:
                            continue
                        out[canal][section].setdefault(day_iso, {'qtd': 0, 'rec': 0.0})
                        out[canal][section][day_iso]['rec'] = v
                elif 'QTD INGRESSOS' in label:
                    for i, day_iso in enumerate(week_dates):
                        q = _parse_qtd(row[i + 1]) if i + 1 < len(row) else 0
                        if q == 0 and day_iso not in out[canal][section]:
                            continue
                        out[canal][section].setdefault(day_iso, {'qtd': 0, 'rec': 0.0})
                        out[canal][section][day_iso]['qtd'] = q
    return out


def build_tm_data_from_planilha(planilha_tm):
    """Constrói ticketmaster-data.json a partir do dict planilha_tm."""
    moto_days = planilha_tm.get('moto', {})
    auto_days = planilha_tm.get('auto', {})
    all_dates = sorted(set(list(moto_days.keys()) + list(auto_days.keys())))

    daily = []
    totals = {'moto_receita': 0.0, 'auto_receita': 0.0, 'moto_ingressos': 0, 'auto_ingressos': 0}
    for ds in all_dates:
        m = moto_days.get(ds, {'qtd': 0, 'rec': 0.0})
        a = auto_days.get(ds, {'qtd': 0, 'rec': 0.0})
        daily.append({
            'date': ds,
            'moto_receita':   round(m['rec'], 2),
            'auto_receita':   round(a['rec'], 2),
            'moto_ingressos': m['qtd'],
            'auto_ingressos': a['qtd'],
        })
        totals['moto_receita']   += m['rec']
        totals['moto_ingressos'] += m['qtd']
        totals['auto_receita']   += a['rec']
        totals['auto_ingressos'] += a['qtd']

    totals['moto_receita']    = round(totals['moto_receita'], 2)
    totals['auto_receita']    = round(totals['auto_receita'], 2)
    totals['total_receita']   = round(totals['moto_receita'] + totals['auto_receita'], 2)
    totals['total_ingressos'] = totals['moto_ingressos'] + totals['auto_ingressos']

    return totals, daily


def build_tm_tipos_from_planilha(planilha_tm):
    """A planilha não tem breakdown por tipo de ingresso (Street Pass, Pista
    Premium, etc.), só Moto vs Auto. Então criamos um único bucket
    'Ticketmaster (planilha)' por canal pra alimentar a seção Tipos.
    Quando a API voltar, granularidade volta automaticamente."""
    def to_bucket(days):
        bucket = {
            'qtd': sum(d['qtd'] for d in days.values()),
            'rec': round(sum(d['rec'] for d in days.values()), 2),
            'daily': {ds: {'qtd': d['qtd'], 'rec': round(d['rec'], 2)}
                      for ds, d in days.items()},
            'por_dia_evento': {},
            'breakdown': {
                'meia':    {'qtd': 0, 'rec': 0.0, 'daily': {}},
                'inteira': {'qtd': 0, 'rec': 0.0, 'daily': {}},
            },
        }
        return bucket

    moto = {'Ticketmaster (planilha)': to_bucket(planilha_tm.get('moto', {}))} if planilha_tm.get('moto') else {}
    auto = {'Ticketmaster (planilha)': to_bucket(planilha_tm.get('auto', {}))} if planilha_tm.get('auto') else {}
    return moto, auto


# ╔══════════════════════════════════════════════════════════╗
# ║  MAIN                                                    ║
# ╚══════════════════════════════════════════════════════════╝

def _load_existing(path):
    """Carrega JSON existente, devolve None se não existir ou inválido."""
    try:
        with open(path, encoding='utf-8') as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return None


_BRT = timezone(timedelta(hours=-3))


def _tm_brt(ts):
    """Timestamp UTC do TM -> (data_brt 'YYYY-MM-DD', hora_brt 0-23). None se invalido."""
    s = (ts or '').strip().replace('Z', '+00:00')
    try:
        dt = datetime.fromisoformat(s)
    except ValueError:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    dt = dt.astimezone(_BRT)
    return dt.strftime('%Y-%m-%d'), dt.hour


def _proprio_brt(cs):
    """created_at 'YYYY-MM-DD HH:MM:SS' (ja BRT) -> (data, hora). None se invalido."""
    s = (cs or '').strip().replace(' ', 'T')
    try:
        dt = datetime.fromisoformat(s)
    except ValueError:
        return None
    return dt.strftime('%Y-%m-%d'), dt.hour


def aggregate_por_hora(tm_movs, moto_sales, auto_sales):
    """Vendas por HORA DO DIA (0-23, BRT), retroativo. Guarda overall (campanha
    toda) E por data, pra o dashboard filtrar por periodo. TM (UTC) -> BRT (-3h);
    Proprio (created_at) ja em BRT. So conta venda REAL (exclui cortesia)."""
    overall = {'moto': [0]*24, 'auto': [0]*24}
    por_data = {}   # 'YYYY-MM-DD' -> {'moto':[24], 'auto':[24]}

    def add(data, hora, edi, q):
        overall[edi][hora] += q
        d = por_data.get(data)
        if d is None:
            d = por_data[data] = {'moto': [0]*24, 'auto': [0]*24}
        d[edi][hora] += q

    for m in (tm_movs or []):
        if m.get('operation') != 'ISSUANCE':
            continue
        if float(m.get('amount', 0) or 0) == 0 or int(m.get('ticketCount', 0) or 0) <= 0:
            continue   # cortesia/vazio nao e venda
        edi = _classify_show(m.get('tickets', []))
        if not edi:
            continue
        r = _tm_brt(m.get('date'))
        if r:
            add(r[0], r[1], edi, int(m.get('ticketCount', 0) or 0))

    for lista, edi in ((moto_sales, 'moto'), (auto_sales, 'auto')):
        for v in (lista or []):
            if str(v.get('venda_status')) not in STATUS_VENDA:
                continue
            r = _proprio_brt(v.get('created_at'))
            if r:
                add(r[0], r[1], edi, len(v.get('qrcodes') or []))

    por_hora = [{'hora': h, 'moto': overall['moto'][h], 'auto': overall['auto'][h],
                 'total': overall['moto'][h]+overall['auto'][h]} for h in range(24)]
    return {
        'updated_at': datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ'),
        'fuso': 'America/Sao_Paulo (BRT)',
        'obs': 'por_hora = campanha toda. por_dia_hora[data] = {moto:[24], auto:[24]} pro filtro de data.',
        'por_hora': por_hora,
        'por_dia_hora': {d: por_data[d] for d in sorted(por_data)},
    }


def main():
    print('=== update-vendas-unificado.py ===')
    print(f'Início: {datetime.now().isoformat()}')

    # ── Sistema Próprio ─ uma única busca, dois agregados ───
    # Se a API do próprio cair, NÃO derruba o resto: o Ticketmaster tem os
    # próprios fallbacks e é ~80% do volume, tem que continuar atualizando.
    # (Em 29/07/2026 a v2.0 quebrou o token e o script morria aqui, no passo 1,
    # levando o Ticketmaster junto — 21h de dado parado no dashboard.)
    proprio_ok, proprio_error = True, None
    moto_sales, auto_sales = [], []
    moto_hist_ok = auto_hist_ok = True
    try:
        print('\n[1/4] Buscando vendas Próprio MOTO...')
        moto_sales, moto_hist_ok = fetch_proprio_sales_resiliente(API_MOTO_BASE, 'MOTO')
        print(f'  {len(moto_sales)} vendas')

        print('[2/4] Buscando vendas Próprio AUTO...')
        auto_sales, auto_hist_ok = fetch_proprio_sales_resiliente(API_AUTO_BASE, 'AUTO')
        print(f'  {len(auto_sales)} vendas')
    except Exception as e:
        proprio_ok, proprio_error = False, str(e)
        moto_sales, auto_sales = [], []
        print(f'  ⚠️  Sistema Próprio falhou: {proprio_error}', file=sys.stderr)

    if proprio_ok:
        diagnostico_recentes(moto_sales, 'MOTO')
        diagnostico_recentes(auto_sales, 'AUTO')
        moto_by_day  = aggregate_proprio_totais(moto_sales)
        auto_by_day  = aggregate_proprio_totais(auto_sales)
        # varredura longa quebrada nao pode apagar o historico nem o dia de hoje
        moto_by_day  = mesclar_historico(moto_by_day, moto_hist_ok,
                                         'moto_receita', 'moto_ingressos', 'MOTO')
        auto_by_day  = mesclar_historico(auto_by_day, auto_hist_ok,
                                         'auto_receita', 'auto_ingressos', 'AUTO')
        proprio_moto = aggregate_proprio_tipos(moto_sales)
        proprio_auto = aggregate_proprio_tipos(auto_sales)
    else:
        # Preserva o último breakdown bom do próprio (0 seria mentira, não falha)
        prev_p       = (_load_existing(OUT_TIPOS) or {}).get('proprio') or {}
        proprio_moto = prev_p.get('moto') or {}
        proprio_auto = prev_p.get('auto') or {}
        moto_by_day  = {}
        auto_by_day  = {}

    # ── Ticketmaster ─ tenta API; se falhar, tenta planilha; senão, mantém antigo ──
    print('\n[3/4] Buscando movimentos Ticketmaster (API)...')
    tm_movs = None
    tm_error = None
    tm_source = None
    tm_cursor_update = None
    tm_cursor_mov_id = None
    try:
        tm_movs, tm_cursor_update, tm_cursor_mov_id = fetch_tm_movements()
        print(f'  {len(tm_movs)} movimentos')
        tm_source = 'api'
    except Exception as e:
        tm_error = str(e)
        print(f'  ⚠️  TM API falhou: {tm_error}', file=sys.stderr)

    if tm_movs is not None:
        tm_totals, tm_daily, tm_last_moto, tm_last_auto = aggregate_tm_totais(tm_movs)
        tm_moto, tm_auto    = aggregate_tm_tipos(tm_movs)
    else:
        # Carrega dados anteriores como baseline (último snapshot bom da API)
        prev_tm    = _load_existing(OUT_TM) or {}
        prev_tipos = _load_existing(OUT_TIPOS) or {}
        prev_daily = {d['date']: d for d in (prev_tm.get('daily') or [])}
        prev_moto  = ((prev_tipos.get('ticketmaster') or {}).get('moto') or {})
        prev_auto  = ((prev_tipos.get('ticketmaster') or {}).get('auto') or {})
        # Sem movimentos frescos pra recalcular ultima venda/cursor: preserva o que ja tinha.
        tm_last_moto = prev_tm.get('last_sale_moto_at')
        tm_last_auto = prev_tm.get('last_sale_auto_at')
        tm_cursor_update = prev_tm.get('tm_cursor_last_update')
        tm_cursor_mov_id = prev_tm.get('tm_cursor_last_movement_id')

        # Fallback 1: planilha. Faz MERGE com o baseline (planilha sobrescreve
        # dias que tem; baseline preserva dias que planilha não cobre — ex.:
        # dias mais recentes que a planilha ainda não foi atualizada manualmente).
        print('  → tentando planilha como fallback...', file=sys.stderr)
        planilha_ok = False
        try:
            planilha = fetch_planilha_data()
            pl_totals, pl_daily = build_tm_data_from_planilha(planilha['tm'])
            # MERGE: parte do prev_daily, sobrescreve dias presentes na planilha
            merged = dict(prev_daily)
            for d in pl_daily:
                merged[d['date']] = d
            tm_daily = sorted(merged.values(), key=lambda x: x['date'])

            # Recalcula totals do merged
            tm_totals = {'moto_receita': 0.0, 'auto_receita': 0.0, 'moto_ingressos': 0, 'auto_ingressos': 0}
            for d in tm_daily:
                tm_totals['moto_receita']   += d.get('moto_receita', 0)
                tm_totals['moto_ingressos'] += d.get('moto_ingressos', 0)
                tm_totals['auto_receita']   += d.get('auto_receita', 0)
                tm_totals['auto_ingressos'] += d.get('auto_ingressos', 0)
            tm_totals['moto_receita']    = round(tm_totals['moto_receita'], 2)
            tm_totals['auto_receita']    = round(tm_totals['auto_receita'], 2)
            tm_totals['total_receita']   = round(tm_totals['moto_receita'] + tm_totals['auto_receita'], 2)
            tm_totals['total_ingressos'] = tm_totals['moto_ingressos'] + tm_totals['auto_ingressos']

            # vendas-tipos: NÃO sobrescreve o breakdown por tipo da API anterior.
            # O front-end já tem reconciliação 'Outros · em sincronização' que
            # absorve a diferença quando o total bate mais alto que a soma das
            # barras. Granularidade volta quando API voltar.
            tm_moto, tm_auto = prev_moto, prev_auto
            print(f'  ✓ Planilha lida: {len(pl_daily)} dias na planilha, '
                  f'merge resultou em {len(tm_daily)} dias totais '
                  f'({tm_totals["moto_ingressos"]} Moto + {tm_totals["auto_ingressos"]} Auto)',
                  file=sys.stderr)
            tm_source = 'planilha'
            planilha_ok = True
        except Exception as pe:
            print(f'  ⚠️  Planilha também falhou: {pe}', file=sys.stderr)

        if not planilha_ok:
            # Fallback 2: só dados anteriores
            tm_totals = prev_tm.get('totals')
            tm_daily  = prev_tm.get('daily', [])
            tm_moto   = prev_moto
            tm_auto   = prev_auto
            tm_source = 'stale'

    tm_ok = tm_source in ('api', 'planilha')

    # ── Escrita dos JSONs ──
    print('\n[4/4] Escrevendo JSONs...')
    if proprio_ok:
        write_json(OUT_VENDAS, build_vendas_data(moto_by_day, auto_by_day))
    else:
        # Mesma regra do TM: não sobrescreve com timestamp mentiroso.
        print(f'  (mantendo {OUT_VENDAS} antigo intacto — API do próprio falhou)')
    if tm_ok:
        write_json(OUT_TM, build_ticketmaster_data(tm_totals, tm_daily, source=tm_source,
                                                    last_sale_moto_at=tm_last_moto,
                                                    last_sale_auto_at=tm_last_auto,
                                                    tm_cursor_last_update=tm_cursor_update,
                                                    tm_cursor_last_movement_id=tm_cursor_mov_id))
    else:
        # NÃO sobrescreve OUT_TM: timestamp ficaria mentiroso. Deixa o arquivo
        # antigo no lugar pra dashboard mostrar 'stale' via updated_at antigo.
        print(f'  (mantendo {OUT_TM} antigo intacto — API e planilha falharam)')
    write_json(OUT_TIPOS, build_tipos_data(proprio_moto, proprio_auto, tm_moto, tm_auto))
    # Vendas por hora do dia (so quando a API TM respondeu, pra nao perder ~80% do volume)
    # Exige os dois lados: com um deles vazio o gráfico por hora ficaria com
    # buraco silencioso (não dá pra distinguir 'não vendeu' de 'não consultei').
    if tm_movs is not None and proprio_ok:
        write_json(OUT_HORA, aggregate_por_hora(tm_movs, moto_sales, auto_sales))
        print(f'  {OUT_HORA} atualizado (vendas por hora do dia)')

    if proprio_ok:
        print(f'\nOK Sistema Próprio atualizado em {OUT_VENDAS} e {OUT_TIPOS} (parte proprio)')
    else:
        print(f'\nSistema Próprio: FALHOU ({proprio_error}) — arquivos antigos preservados')
        # Anotação do Actions: o run continua verde (pra não travar o commit do
        # Ticketmaster), mas o aviso aparece no resumo da execução.
        print(f'::warning title=Sistema Proprio sem atualizar::{proprio_error}')
    print(f'Ticketmaster: source={tm_source}')
    if tm_source == 'stale':
        print(f'  Causa API: {tm_error}')
    print(f'Fim: {datetime.now().isoformat()}')


if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print(f'\nERRO: {e}', file=sys.stderr)
        sys.exit(1)
